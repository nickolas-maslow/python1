import psycopg2
import openpyxl
from openpyxl.utils import get_column_letter

connection = psycopg2.connect(host='localhost',
                              user='postgres',
                              password='1111',
                              database='new_postgres_bd'
                              )
cursor = connection.cursor()
cursor.execute("""
SELECT * FROM public.new_postgres_tb
WHERE id <= 50
"""
)
records = cursor.fetchall()
connection.close()

book = openpyxl.Workbook()
sheet = book.active

titles = ['user 1', 'user 2', 'user 3']
index_col = 1
for i in titles:
    sheet[f'{get_column_letter(index_col)}1'] = str(i)
    index_col += 1

index_row = 2
for i in records:
    col_index = 1
    for column in i:
        sheet[f'{get_column_letter(col_index)}{index_row}'] = str(column)
        col_index += 1
    index_row += 1
book.save('new_table_sql.xlsx')
