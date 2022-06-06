import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


workbook = openpyxl.load_workbook("""sample.xlsx""")

worksheet = workbook.active

max_row = worksheet.max_row
print(max_row)
max_column = worksheet.max_column
print(max_column)

external_array = []

for row in range(1, max_row + 1):
    internal_array = []

    for col in range(1, max_column + 1):
        value = worksheet.cell(row=row, column=col).value
        if value is None:
            value = ''
            continue
            break
        internal_array.append(value)
    # завершение наполнение внутреннего массива

    if len(internal_array) < 1:
        continue

    external_array_array.append(internal_array)

print(external_array)

index = 0
while True:
    index += 1
    if index > 50:
        break

workbook_new = Workbook()
worksheet_new = workbook_new.active

col_count = external_array[0]
print(
    f"col_count: {external_array}")
print(f"col_count: {external_array[0]}")
print(f"col_count: {external_array[0][1]}")
print(f"col_count: {external_array[0][1][2:-2:1]}")


def function_len_array(array):
    return len(array)


for row in range(0, function_len_array(external_array)):
    print(f"col_count: {len(external_array[row])}")
    for col in range(0, function_len_array(external_array[row])):
        worksheet[f'{get_column_letter(col + 1)}{row + 1}'] = external_array[row][col]
        if row == 0:
            worksheet[f'{get_column_letter(col + 1)}{row + 1}'].font = Font(bold=True)
        pass
    print(external_array[row])
    pass

for char in 'ABCDE':
    worksheet_new['A1'].font = Font(bold=True)
    worksheet_new['B1'].font = Font(bold=True)
    worksheet_new['C1'].font = Font(bold=True)
    worksheet_new['D1'].font = Font(bold=True)
    worksheet_new['E1'].font = Font(bold=True)


workbook.save('NEW_SAMPLE.xlsx')

wb = Workbook()
ws = wb.active
ws['B2'] = "Hello1"
ws['B3'] = "Hello"
ws['B2'].font = Font(bold=True)
ws['B3'].font = Font(bold=True)
wb.save("demo.xlsx")