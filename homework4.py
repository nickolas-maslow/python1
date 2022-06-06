import openpyxl
from exaple_list import get
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

book1 = Workbook()
page1 = book1.active

var_range = range(1, 26)
line_1 = []
line1 = ""
for i in var_range:
    line1 = f"_{i}"
    line_1.append(line1)

line_2 = []
line = ""
for j in var_range:
    line2 = f"A_{j}"
    line_2.append(line2)

final_line = []
final_line.append(line_1)
final_line.append(line_2)
print(final_line)

for row in range(0, len(final_line)):
    for col in range(0, len(final_line[row])):
        print(final_line[row])
        col_letter = get_column_letter(col)
        page1[f"{get_column_letter(col + 1)}{row + 1}"] = final_line[row][col]

for col in range(0, len(final_line)):
    for row in range(0, len(final_line[col])):
        page1.cell(row=row + 4, column=col + 1).value = final_line[col][row]
print(final_line)

book1.save("home_work4.xlsx")