import openpyxl
from openpyxl import Workbook

class MyCalculator:
    @staticmethod
    def plus(val1: float, val2: float):
        return val1 + val2

    def minus(val1: float, val2: float):
        return val1 - val2

    def multiply(val1: float, val2: float):
        return val1 * val2

    def divide(val1: float, val2: float):
        if val2 != 0:
            return val1 / val2
        else:
            print("Error")
            return 0


MyCalculator.plus(16, 24)
print(MyCalculator.plus(16, 24))


##############

class GenerateMassive:
    def __init__(self, name: str, index: str, value: str):
        self.name = name
        self.index = index
        self.value = value

    def get_name(self):
        return f"{self.name} {self.index} {self.value}"

run = []
for i in range(1, 1001):
    run.append(GenerateMassive(name=f"_{i}", index=f"{i}", value=f"_A{i}").get_name())
print(run)


#############


from openpyxl.utils import get_column_letter

file_name = "file.xlsx"
file_locate = "folder/file.xlsx"
workbook = openpyxl.load_workbook(file_locate + "/" + file_name)
worksheet = workbook.active
max_row = worksheet.max_row
max_column = worksheet.max_column


class Vacation_first_part:
    def __init__(self, worker: str, year_start: int, start_date: int, day_duration: str, end_date: int):
        self.worker = worker
        self.year_start = year_start
        self.start_date = start_date
        self.day_duration = day_duration
        self.end_date = end_date

    def info(self):
        return f"{self.worker}{self.year_start}{self.start_date}{self.day_duration}{self.end_date}"


massive = []
for row in range(1, max_row+1):
    add_info = []
    for column in range(1, max_column+1):
        obj = worksheet.cell(row=row, column=column).value
        add_info.append(obj)
        name = Vacation_first_part(
            worker=add_info[0],
            year_start=add_info[1],
            start_date=add_info[2],
            day_duration=[3],
            end_date=[4]
        )
        print(name)

book = Workbook()
page = book.active
for row in range(0, len(massive)):
    for col in range(0, len(massive[row])):
        page.cell(row=col + 1, column=row + 1).value = massive[row][column]

var = 0

for row in range(0, len(massive)):
    if massive[row].is_pribyl():
        var += 1
        page[f"A{var + 1}"] = massive[row].worker
        page[f"B{var + 1}"] = massive[row].year_start
        page[f"B{var + 1}"] = massive[row].start_date
        page[f"B{var + 1}"] = massive[row].day_duration
        page[f"B{var + 1}"] = massive[row].end_date
        for col in range(0, len(massive[row])):
            print([row])
            letter = get_column_letter(column)
            page[f"{get_column_letter(col + 1)}{var + 1}"] = massive[row][col]

book.save("file.xlsx")
