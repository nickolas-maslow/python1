from openpyxl import Workbook
from openpyxl.utils import get_column_letter

workbook = Workbook()

worksheet = workbook.active

var_list = ["Zhezkazgan", "Karaganda", "Taraz", "Almaty"]

# ["Zhezkazgan", "Karaganda", "Taraz", "Almaty"]
# ["Zhezkazgan", "Karaganda", "Taraz", "Almaty"]
# ["Zhezkazgan", "Karaganda", "Taraz", "Almaty"]
# ["Zhezkazgan", "Karaganda", "Taraz", "Almaty"]

var_range = range(1, 1000, 1)
print(var_range)
for n in var_range:
    print(n)

x = range(6+1)
for n in x:
    print(n)
pass

x = range(3, 200, 2)
for n in x:
    print(n)

for i in "ABCD":
    row = "1"
    col = i

worksheet[f'{col}{row}'] = str(var_list["ABCD".index(i)])

for number in var_range:
    for name in var_list:
        row = number

        col = get_column_letter(var_list.index(name) + 1)

        worksheet[f'{col}{row}'] = str(name)

worksheet['C!'] = 11

# сохраняем рабочую книгу в excel
workbook.save("sample.xlsx")