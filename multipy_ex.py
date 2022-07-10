import openpyxl

workbook = openpyxl.load_workbook("homework_dir/data.xlsx")
worksheet = workbook.active

max_row1 = worksheet.max_row + 1
max_column1 = worksheet.max_column + 1

out_list_ = []
for row in range(1, max_row1 + 1):
    in_list_ = []
    for col in range(1, max_column1 + 1):
        value = worksheet.cell(row=row, column=col).value
        if value is None:
            value = ""
        in_list_.append(value)
    out_list_.append(in_list_)
print(out_list_)

line1 = set(out_list_[1])
line1.remove("")
print(line1)
line2 = set(out_list_[3])
line2.remove("")
print(line2)
line3 = set(out_list_[5])
line3.remove("")
print(line3)

line4 = line2.intersection(line1, line3)
print(line4)

line5 = line2.difference(line1, line3)
print(line5)

line6 = line3.union(line1)
print(line6)
