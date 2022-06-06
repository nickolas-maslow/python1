import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval

name_of_file = "sample.xlsx"

workbook = openpyxl.load_workbook(name_of_file)

worksheet = workbook.active

max_row = worksheet.max_row + 1
print(max_row)
print(type(max_row))

max_column = worksheet.max_column + 1
print(max_column)
print(type(max_column))

index = 0
for i in range(1, max_row):
    value = worksheet.cell(row=i, column=1).value
    value = worksheet[f"A{i}"].value

    if value is not None:
        pass

    if len(str(value)) >= 2:
        if value:
            print(value)
            print(type(value))
            index += 1

    for j in range(1, max_column):
        value = worksheet.cell(row=i, column=j).value
        print(f"index: {i} {j}")

    for j in range(1, max_column):
        value = worksheet.cell(row=i, column=j).value
        print(f"index: {i} {j}")

        value = worksheet[f"A{i}"].value

        "Karaganda Shymkent"

        if value is not None:
            pass
        if len(str(value)) >= 2:
            if (value == "Almaty") or (value == "Taraz"):
                print(value)
                print(type(value))
                index += 1

print(index)
    for i in range(1, max_row):
        for j in "ABCD":
            for name in var_list:
                row = num
                col = get_column_letter(var_list.index(name) + 1)
                worksheet[f'{col}{row}'] = str(name)

yes = True
no = False

if (yes) and not (no):
    print("Agreement")
else:
    print('Disagreement')


