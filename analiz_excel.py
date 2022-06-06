from tkinter import *
from tkinter import ttk
import tkinter
from tkinter import messagebox
import openpyxl


def cancel():
    pass


def get_result():
    col_from_user = col_label.get()
    print(f'колонки: {col_from_user}')
    elem_from_user = elem_label.get()
    print(f'элемент: {elem_from_user}')

    name_of_file = 'sample.xlsx'
    workbook = openpyxl.load_workbook(name_of_file)

    worksheet = workbook.active

    max_row = worksheet.max_row
    print(max_row)
    max_column = worksheet.max_column
    print(max_column)

    extended_array = []
    for row in range(1, max_row + 1):
        internal_array = []
        for col in range(1, max_column + 1):
            value = worksheet.cell(row=row, column=col).value
            if value is None:
                value = ''
        internal_array.append(value)

        if len(internal_array) < 1:
            continue

        extended_array.append(internal_array)

    print(extended_array)

    col_list = []
    print(col_from_user)
    print(type(col_from_user))
    for i in col_from_user.split(","):
        i = str(i).strip()
        i = int(i)  # "1" -> 1
        col_list.append(i)
        print(i)
        print(type(i))
    print(col_list)

    count = 0
    for i_new in extended_array:
        for j_new in i_new:
            if elem_from_user == j_new:
                count = count + 1

    result = count
    result_label.config(text=f"{result}")


root = Tk()

frm = ttk.Frame(root, padding=100)
root.title("Анализ данных")
root.geometry("640x480")
frm.grid()

result_label = ttk.Label(frm, text="Количество вхождений:")
result_label.grid(column=4, row=0)

col_label_var = tkinter.StringVar()
col_label = ttk.Entry(textvariable=col_label_var)
col_label.grid(column=1, row=3)

elem_label_var = tkinter.StringVar()  # переменная, которая хранит элемент
elem_label = ttk.Entry(textvariable=elem_label_var)
elem_label.grid(column=5, row=3)

Button(text="stop",
       background="#555",
       foreground="#ccc",
       padx="20",
       pady="8",
       font="16",
       command=cancel,
       ).grid(column=0, row=5)

Button(text="start",
       background="#555",
       foreground="#ccc",
       padx="20",
       pady="8",
       font="16",
       command=get_result,
       ).grid(column=2, row=5)

root.mainloop()

str1 = "124124|14121|1411234".split(sep="|")